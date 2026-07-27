"""
Interfaz grafica para calcular la Curva de Duracion de Caudales (FDC)
======================================================================
Permite seleccionar un archivo de estacion (.xlsx), elegir mediante una
ventana emergente que celda contiene el nombre de la estacion y cuales
celdas son el encabezado de la columna de fechas y de la columna de
caudales (el formato de los archivos de entrada puede variar), luego
elegir los anos a incluir en el analisis y calcular los percentiles
Q2.75, Q50 y Q97.25 sobre los datos diarios seleccionados.

Formula de la curva de duracion (Weibull):
    p_i = 100 * i / (n + 1)

Uso:
    python fdc_gui.py
"""

import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

P_ALTO = 2.75
P_MED = 50.0
P_BAJO = 97.25

MAX_PREVIEW_ROWS = 60
MAX_PREVIEW_COLS = 15


class CellPickerDialog(tk.Toplevel):
    """
    Ventana emergente que muestra una vista previa de la hoja de calculo
    y permite al usuario elegir, haciendo clic en la grilla:
      - la celda con el nombre de la estacion
      - la celda de encabezado de la columna de fechas
      - la celda de encabezado de la columna de caudales

    Resultado en self.result (dict) con claves "nombre", "fecha", "valor",
    cada una como tupla (fila, columna) en base 1, o None si se cancela.
    """

    MODOS = [
        ("nombre", "Nombre de la estacion"),
        ("fecha", "Encabezado columna de Fechas"),
        ("valor", "Encabezado columna de Caudales"),
    ]

    def __init__(self, parent, ws):
        super().__init__(parent)
        self.title("Seleccionar celdas de referencia")
        self.geometry("900x600")
        self.transient(parent)
        self.grab_set()

        self.result = None
        self.selecciones = {}
        self.modo = tk.StringVar(value="nombre")

        n_filas = min(ws.max_row or MAX_PREVIEW_ROWS, MAX_PREVIEW_ROWS)
        n_cols = min(ws.max_column or MAX_PREVIEW_COLS, MAX_PREVIEW_COLS)
        self._filas_preview = list(ws.iter_rows(
            min_row=1, max_row=n_filas, max_col=n_cols, values_only=True))
        self.n_cols = n_cols

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=8)
        top.pack(side="top", fill="x")
        ttk.Label(top, text="Modo de seleccion:").pack(side="left")
        for valor, etiqueta in self.MODOS:
            ttk.Radiobutton(top, text=etiqueta, value=valor,
                             variable=self.modo).pack(side="left", padx=6)

        info = ttk.Frame(self, padding=(8, 0))
        info.pack(side="top", fill="x")
        self.lbl_sel = ttk.Label(info, text=self._texto_seleccion())
        self.lbl_sel.pack(side="left")

        ttk.Label(self, text="Haga clic en la celda correspondiente en la tabla:",
                  padding=(8, 4)).pack(side="top", anchor="w")

        grid_frame = ttk.Frame(self)
        grid_frame.pack(side="top", fill="both", expand=True, padx=8, pady=4)

        cols = [str(i) for i in range(1, self.n_cols + 1)]
        self.tree = ttk.Treeview(grid_frame, columns=cols, show="headings",
                                  height=20)
        for i, c in enumerate(cols):
            letra = get_column_letter(i + 1)
            self.tree.heading(c, text=letra)
            self.tree.column(c, width=80, anchor="w")

        vsb = ttk.Scrollbar(grid_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(grid_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        grid_frame.rowconfigure(0, weight=1)
        grid_frame.columnconfigure(0, weight=1)

        for fila_idx, fila in enumerate(self._filas_preview, start=1):
            valores = ["" if v is None else str(v) for v in fila]
            self.tree.insert("", "end", iid=str(fila_idx),
                              text=str(fila_idx), values=valores)

        self.tree.bind("<Button-1>", self._on_click)

        botones = ttk.Frame(self, padding=8)
        botones.pack(side="bottom", fill="x")
        ttk.Button(botones, text="Cancelar", command=self._cancelar).pack(side="right", padx=4)
        ttk.Button(botones, text="Aceptar", command=self._aceptar).pack(side="right", padx=4)

    def _texto_seleccion(self):
        partes = []
        for valor, etiqueta in self.MODOS:
            celda = self.selecciones.get(valor)
            if celda:
                letra = get_column_letter(celda[1])
                partes.append(f"{etiqueta}: {letra}{celda[0]}")
            else:
                partes.append(f"{etiqueta}: (sin elegir)")
        return "   |   ".join(partes)

    def _on_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        fila_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not fila_id or not col_id:
            return
        fila = int(fila_id)
        col = int(col_id.replace("#", ""))

        self.selecciones[self.modo.get()] = (fila, col)
        self.lbl_sel.config(text=self._texto_seleccion())

    def _aceptar(self):
        faltantes = [etq for val, etq in self.MODOS if val not in self.selecciones]
        if faltantes:
            messagebox.showwarning(
                "Seleccion incompleta",
                "Falta seleccionar:\n- " + "\n- ".join(faltantes),
                parent=self)
            return
        self.result = dict(self.selecciones)
        self.destroy()

    def _cancelar(self):
        self.result = None
        self.destroy()


def cargar_datos(ruta, celda_nombre, celda_fecha, celda_valor):
    """
    Lee nombre de estacion y serie Fecha/Valor a partir de celdas de
    referencia elegidas por el usuario.

    celda_nombre: (fila, columna) de la celda con el nombre de la estacion.
    celda_fecha:  (fila, columna) del encabezado de la columna de fechas;
                  los datos se leen desde la fila siguiente hacia abajo.
    celda_valor:  (fila, columna) del encabezado de la columna de caudales;
                  los datos se leen desde la fila siguiente hacia abajo.
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[wb.sheetnames[0]]

    nombre_estacion = ws.cell(row=celda_nombre[0], column=celda_nombre[1]).value
    nombre_estacion = str(nombre_estacion).strip() if nombre_estacion else os.path.basename(ruta)

    col_fecha = celda_fecha[1]
    col_valor = celda_valor[1]
    fila_inicio = max(celda_fecha[0], celda_valor[0]) + 1

    fechas, valores = [], []
    fila = fila_inicio
    while True:
        f = ws.cell(row=fila, column=col_fecha).value
        v = ws.cell(row=fila, column=col_valor).value
        if f is None and v is None:
            break
        fechas.append(f)
        valores.append(v)
        fila += 1

    df = pd.DataFrame({"Fecha": fechas, "Valor": valores})
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
    df = df.dropna(subset=["Fecha", "Valor"]).sort_values("Fecha").reset_index(drop=True)
    return nombre_estacion, df


def calcular_fdc(serie):
    datos = np.sort(serie.values)[::-1]
    n = len(datos)
    exceed = 100.0 * np.arange(1, n + 1) / (n + 1)
    return pd.DataFrame({"excedencia": exceed, "caudal": datos})


def percentil_fdc(fdc_df, prob_excedencia):
    return float(np.interp(prob_excedencia, fdc_df["excedencia"], fdc_df["caudal"]))


class FDCApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Curva de Duracion de Caudales (FDC)")
        self.geometry("1200x750")

        self.nombre_estacion = None
        self.df = None
        self.ruta_actual = None

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=8)
        top.pack(side="top", fill="x")

        ttk.Button(top, text="Abrir archivo...", command=self.abrir_archivo).pack(side="left")
        self.lbl_archivo = ttk.Label(top, text="Sin archivo cargado")
        self.lbl_archivo.pack(side="left", padx=10)

        body = ttk.Frame(self)
        body.pack(side="top", fill="both", expand=True)

        # Panel izquierdo: seleccion de anos
        left = ttk.Frame(body, padding=8)
        left.pack(side="left", fill="y")

        ttk.Label(left, text="Anos disponibles:").pack(anchor="w")

        listbox_frame = ttk.Frame(left)
        listbox_frame.pack(fill="y", expand=True)
        scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical")
        self.listbox = tk.Listbox(listbox_frame, selectmode="extended", height=25,
                                   exportselection=False, yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listbox.yview)
        self.listbox.pack(side="left", fill="y")
        scrollbar.pack(side="left", fill="y")

        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=6)
        ttk.Button(btns, text="Todos", command=self.seleccionar_todos).pack(side="left", padx=2)
        ttk.Button(btns, text="Ninguno", command=self.seleccionar_ninguno).pack(side="left", padx=2)

        ttk.Button(left, text="Calcular FDC", command=self.calcular,
                   style="Accent.TButton").pack(fill="x", pady=10)

        self.txt_resultados = tk.Text(left, width=32, height=10, state="disabled")
        self.txt_resultados.pack(fill="x", pady=6)

        # Panel derecho: grafica
        right = ttk.Frame(body, padding=8)
        right.pack(side="left", fill="both", expand=True)

        self.fig, self.ax = plt.subplots(figsize=(9, 6))
        self.canvas = FigureCanvasTkAgg(self.fig, master=right)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def abrir_archivo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo de estacion",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=os.path.dirname(os.path.abspath(__file__)),
        )
        if not ruta:
            return

        try:
            wb_preview = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
            ws_preview = wb_preview[wb_preview.sheetnames[0]]
            dialog = CellPickerDialog(self, ws_preview)
            self.wait_window(dialog)
            wb_preview.close()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo previsualizar el archivo:\n{e}")
            return

        if dialog.result is None:
            return

        try:
            nombre, df = cargar_datos(ruta, dialog.result["nombre"],
                                       dialog.result["fecha"], dialog.result["valor"])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")
            return

        if df.empty:
            messagebox.showwarning("Aviso", "No se encontraron datos validos en el archivo.")
            return

        self.ruta_actual = ruta
        self.nombre_estacion = nombre
        self.df = df
        self.lbl_archivo.config(text=f"{os.path.basename(ruta)}  —  Estacion: {nombre}")

        anos = sorted(df["Fecha"].dt.year.unique())
        self.listbox.delete(0, tk.END)
        for a in anos:
            self.listbox.insert(tk.END, str(a))
        self.listbox.select_set(0, tk.END)  # todos seleccionados por defecto

    def seleccionar_todos(self):
        self.listbox.select_set(0, tk.END)

    def seleccionar_ninguno(self):
        self.listbox.select_clear(0, tk.END)

    def calcular(self):
        if self.df is None:
            messagebox.showwarning("Aviso", "Primero abra un archivo de datos.")
            return

        seleccion = [int(self.listbox.get(i)) for i in self.listbox.curselection()]
        if not seleccion:
            messagebox.showwarning("Aviso", "Seleccione al menos un ano.")
            return

        df_sel = self.df[self.df["Fecha"].dt.year.isin(seleccion)]
        serie = df_sel["Valor"].dropna()
        if serie.empty:
            messagebox.showwarning("Aviso", "No hay datos para los anos seleccionados.")
            return

        fdc = calcular_fdc(serie)
        q275 = percentil_fdc(fdc, P_ALTO)
        q50 = percentil_fdc(fdc, P_MED)
        q9725 = percentil_fdc(fdc, P_BAJO)
        maximos_anuales = df_sel.groupby(df_sel["Fecha"].dt.year)["Valor"].max()
        minimos_anuales = df_sel.groupby(df_sel["Fecha"].dt.year)["Valor"].min()
        q_max = float(maximos_anuales.mean())
        q_min = float(minimos_anuales.mean())
        q_prom = float(serie.mean())

        q275 = min(q275, float(serie.max()))
        q9725 = max(q9725, float(serie.min()))

        self._mostrar_resultados(seleccion, len(serie), q275, q50, q9725, q_max, q_min, q_prom)
        self._graficar(fdc, seleccion, q275, q50, q9725, q_max, q_min, q_prom)

    def _mostrar_resultados(self, anos, n, q275, q50, q9725, q_max, q_min, q_prom):
        anos_txt = self._formatear_anos(anos)
        texto = (
            f"Estacion: {self.nombre_estacion}\n"
            f"Anos: {anos_txt}\n"
            f"N datos: {n}\n\n"
            f"Q2.75  = {q275:.3f} m3/s\n"
            f"Q50    = {q50:.3f} m3/s\n"
            f"Q97.25 = {q9725:.3f} m3/s\n\n"
            f"Q_max (prom. anual)  = {q_max:.3f} m3/s\n"
            f"Q_min (prom. anual)  = {q_min:.3f} m3/s\n"
            f"Q_prom               = {q_prom:.3f} m3/s\n"
        )
        self.txt_resultados.config(state="normal")
        self.txt_resultados.delete("1.0", tk.END)
        self.txt_resultados.insert(tk.END, texto)
        self.txt_resultados.config(state="disabled")

    @staticmethod
    def _formatear_anos(anos):
        anos = sorted(anos)
        if len(anos) > 8:
            return f"{anos[0]}-{anos[-1]} ({len(anos)} anos)"
        return ", ".join(str(a) for a in anos)

    def _graficar(self, fdc, anos, q275, q50, q9725, q_max, q_min, q_prom):
        ax = self.ax
        ax.clear()

        caudales_pos = fdc["caudal"].clip(lower=1e-6)
        ax.semilogy(fdc["excedencia"], caudales_pos, color="#1f77b4",
                    linewidth=2.0, label=f"FDC — {self.nombre_estacion}", zorder=3)

        marcadores = {
            P_ALTO: ("^", "Q2.75", q275),
            P_MED: ("o", "Q50", q50),
            P_BAJO: ("v", "Q97.25", q9725),
        }
        for p_exc, (mk, etiq, val) in marcadores.items():
            ax.plot(p_exc, val, marker=mk, color="#1f77b4", markersize=9,
                    zorder=5, linestyle="None",
                    label=f"{etiq} = {val:.3f} m³/s")
            ax.axvline(x=p_exc, color="dimgrey", linestyle="--",
                       linewidth=0.9, alpha=0.55, zorder=1)

        ylim = ax.get_ylim()
        y_top = ylim[1] * 0.9
        for p_exc, (_, etiq, _) in marcadores.items():
            ax.text(p_exc + 0.5, y_top, etiq, color="dimgrey", fontsize=8,
                    va="top", fontstyle="italic")

        ax.set_xlim(0, 100)
        ax.grid(True, which="both", linestyle="--", alpha=0.3)
        ax.set_xlabel("Probabilidad de excedencia (%)", fontsize=11)
        ax.set_ylabel("Caudal (m³/s)  [escala log]", fontsize=11)

        anos_txt = self._formatear_anos(anos)
        ax.set_title(f"{self.nombre_estacion}  —  FDC  —  Anos: {anos_txt}",
                     fontsize=12, fontweight="bold", pad=10)
        ax.legend(fontsize=8.5, loc="upper right", framealpha=0.9)

        info = (f"Q_max (prom. anual) = {q_max:.3f}  |  "
                f"Q_min (prom. anual) = {q_min:.3f}  |  "
                f"Q_prom = {q_prom:.3f}  m³/s")
        ax.text(0.5, -0.12, info, transform=ax.transAxes, ha="center",
                fontsize=9, color="grey")

        self.fig.tight_layout()
        self.canvas.draw()


if __name__ == "__main__":
    app = FDCApp()
    app.mainloop()
